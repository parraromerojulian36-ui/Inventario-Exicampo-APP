import streamlit as st
import pandas as pd
import openpyxl
import os
import tempfile
from datetime import datetime
from PIL import Image
import pytesseract
import cv2
import numpy as np
import re

# ==========================================
# CONFIGURACIÓN UI
# ==========================================
st.set_page_config(
    page_title="PASCA Inventory Audit Pro",
    layout="wide"
)

st.markdown("""
<style>

.stNumberInput label {
    font-size:16px !important;
    font-weight:bold !important;
}

.big-font {
    font-size:32px;
    font-weight:bold;
    text-align:center;
    color:white;
    background:#2E7D32;
    padding:15px;
    border-radius:15px;
    margin-top:20px;
}

.product-header {
    background:white;
    padding:20px;
    border-radius:15px;
    border-left:10px solid #4CAF50;
    margin-bottom:20px;
}

</style>
""", unsafe_allow_html=True)

# ==========================================
# UTILIDADES
# ==========================================
def clean_code(val):

    if pd.isna(val):
        return ""

    val = str(val).strip()

    if val.endswith(".0"):
        val = val[:-2]

    return val


# ==========================================
# PREPROCESAMIENTO OCR
# ==========================================
def preprocess_image(image):

    img = np.array(image)

    gray = cv2.cvtColor(
        img,
        cv2.COLOR_RGB2GRAY
    )

    gray = cv2.GaussianBlur(
        gray,
        (3, 3),
        0
    )

    thresh = cv2.threshold(
        gray,
        0,
        255,
        cv2.THRESH_BINARY + cv2.THRESH_OTSU
    )[1]

    return thresh


# ==========================================
# OCR LOCAL
# ==========================================
def identify_product_local(image, df_sistema):

    try:

        processed = preprocess_image(image)

        text = pytesseract.image_to_string(
            processed
        )

        text = text.upper().strip()

        st.info(f"OCR detectado: {text}")

        # ======================================
        # BUSCAR CODIGO
        # ======================================
        codes = (
            df_sistema.iloc[:, 0]
            .astype(str)
            .tolist()
        )

        best_code = None
        best_score = 0

        for code in codes:

            if code in text:
                best_code = code
                best_score = 100
                break

        if best_code:

            row = df_sistema[
                df_sistema.iloc[:, 0]
                .astype(str) == best_code
            ]

            if not row.empty:

                return {
                    "code": best_code,
                    "name": row.iloc[0, 1]
                }

        # ======================================
        # BUSCAR NOMBRE
        # ======================================
        names = (
            df_sistema.iloc[:, 1]
            .astype(str)
            .tolist()
        )

        best_name = None
        best_words = 0

        for name in names:

            words = name.upper().split()

            score = sum(
                1 for w in words
                if w in text
            )

            if score > best_words:
                best_words = score
                best_name = name

        if best_name and best_words >= 1:

            row = df_sistema[
                df_sistema.iloc[:, 1]
                .astype(str) == best_name
            ]

            if not row.empty:

                return {
                    "code": clean_code(row.iloc[0, 0]),
                    "name": best_name
                }

        return None

    except Exception as e:

        st.error(str(e))
        return None


# ==========================================
# AGREGAR PRODUCTO
# ==========================================
def add_product_to_conteo(
    df_conteo,
    code,
    name
):

    exists = df_conteo[
        df_conteo.iloc[:, 0]
        .astype(str) == code
    ]

    if exists.empty:

        new_row = (
            [code, name] +
            [0] * (len(df_conteo.columns) - 2)
        )

        df_conteo.loc[
            len(df_conteo)
        ] = new_row

    return df_conteo


# ==========================================
# CARGAR EXCEL
# ==========================================
def load_pasca_data(uploaded_file):

    with tempfile.NamedTemporaryFile(
        delete=False,
        suffix=".xlsx"
    ) as tmp:

        tmp.write(
            uploaded_file.getvalue()
        )

        tmp_path = tmp.name

    wb = openpyxl.load_workbook(
        tmp_path
    )

    # ======================================
    # SISTEMA
    # ======================================
    df_sistema = pd.read_excel(
        tmp_path,
        sheet_name="SISTEMA"
    )

    df_sistema.columns = (
        df_sistema.columns
        .str.strip()
    )

    df_sistema.iloc[:, 0] = (
        df_sistema.iloc[:, 0]
        .apply(clean_code)
    )

    # ======================================
    # CONTEO
    # ======================================
    df_conteo = pd.read_excel(
        tmp_path,
        sheet_name="CONTEO_F"
    )

    header_row = 0

    for i, row in df_conteo.iterrows():

        if "CODIGO" in str(
            row.values
        ).upper():

            header_row = i
            break

    df_conteo.columns = (
        df_conteo.iloc[header_row]
        .astype(str)
        .str.strip()
    )

    df_conteo = (
        df_conteo.iloc[header_row + 1:]
        .reset_index(drop=True)
    )

    df_conteo = df_conteo.astype(object)

    df_conteo.iloc[:, 0] = (
        df_conteo.iloc[:, 0]
        .apply(clean_code)
    )

    st.session_state.temp_file = tmp_path

    return df_conteo, df_sistema, wb


# ==========================================
# EXPORTAR
# ==========================================
def save_full_audit(
    df_conteo,
    df_sistema,
    wb
):

    # ======================================
    # CONTEO_F
    # ======================================
    sheet = wb["CONTEO_F"]

    start_row = 1

    for row in sheet.iter_rows(
        max_row=10
    ):

        for cell in row:

            if (
                cell.value and
                "CODIGO" in str(
                    cell.value
                ).upper()
            ):

                start_row = (
                    cell.row + 1
                )

                break

    # escribir
    for i, row in df_conteo.iterrows():

        row_num = start_row + i

        for col_num, value in enumerate(
            row.values,
            1
        ):

            sheet.cell(
                row=row_num,
                column=col_num
            ).value = value

    # ======================================
    # RESULTADO
    # ======================================
    sheet_res = wb["RESULTADO"]

    for row in sheet_res.iter_rows(
        min_row=5
    ):

        for cell in row:
            cell.value = None

    row_res = 5

    for _, row_c in df_conteo.iterrows():

        code = clean_code(
            row_c.iloc[0]
        )

        name = row_c.iloc[1]

        total_fisico = (
            row_c.iloc[11]
            if pd.notnull(
                row_c.iloc[11]
            )
            else 0
        )

        match = df_sistema[
            df_sistema.iloc[:, 0]
            .astype(str) == code
        ]

        if not match.empty:

            total_sistema = (
                match.iloc[0, 2]
                if pd.notnull(
                    match.iloc[0, 2]
                )
                else 0
            )

            diff = (
                total_fisico -
                total_sistema
            )

            faltante = (
                abs(diff)
                if diff < 0
                else 0
            )

            sobrante = (
                diff
                if diff > 0
                else 0
            )

            sheet_res.cell(
                row=row_res,
                column=1
            ).value = code

            sheet_res.cell(
                row=row_res,
                column=2
            ).value = name

            sheet_res.cell(
                row=row_res,
                column=3
            ).value = total_fisico

            sheet_res.cell(
                row=row_res,
                column=4
            ).value = total_sistema

            sheet_res.cell(
                row=row_res,
                column=5
            ).value = diff

            sheet_res.cell(
                row=row_res,
                column=6
            ).value = faltante

            sheet_res.cell(
                row=row_res,
                column=7
            ).value = sobrante

            row_res += 1

    with tempfile.NamedTemporaryFile(
        delete=False,
        suffix=".xlsx"
    ) as tmp:

        path = tmp.name

    wb.save(path)

    with open(path, "rb") as f:
        data = f.read()

    try:

        os.remove(
            st.session_state.temp_file
        )

        os.remove(path)

    except:
        pass

    return data


# ==========================================
# UI
# ==========================================
st.title(
    "📦 PASCA Inventory Audit Pro"
)

with st.sidebar:

    sucursal = st.selectbox(
        "Sucursal",
        [
            "PASCA",
            "SUBIA",
            "SIBATE",
            "GRANADA"
        ]
    )

    fecha = datetime.now().strftime(
        "%d-%m-%Y"
    )

uploaded_file = st.file_uploader(
    "Sube Excel",
    type=["xlsx"]
)

# ==========================================
# APP
# ==========================================
if uploaded_file:

    # ======================================
    # SESSION
    # ======================================
    if "df_inv" not in st.session_state:

        df_c, df_s, wb = (
            load_pasca_data(
                uploaded_file
            )
        )

        st.session_state.df_inv = df_c
        st.session_state.df_sistema = df_s
        st.session_state.wb = wb

    df_conteo = st.session_state.df_inv
    df_sistema = st.session_state.df_sistema
    wb = st.session_state.wb

    # ======================================
    # CAMARA
    # ======================================
    st.subheader("📷 Cámara")

    img_file = st.camera_input(
        "Tomar foto producto"
    )

    if img_file:

        img = Image.open(img_file)

        with st.spinner(
            "Analizando..."
        ):

            result = identify_product_local(
                img,
                df_sistema
            )

        if result:

            code = result["code"]
            name = result["name"]

            st.success(
                f"Detectado: {name}"
            )

            st.session_state.df_inv = (
                add_product_to_conteo(
                    df_conteo,
                    code,
                    name
                )
            )

            st.session_state.selected_code = code
            st.session_state.selected_name = name

            st.rerun()

        else:

            st.error(
                "No se encontró coincidencia"
            )

    # ======================================
    # BUSQUEDA MANUAL
    # ======================================
    st.subheader("🔍 Buscar")

    search = st.text_input(
        "Código o nombre"
    ).upper()

    if search:

        mask = (
            (
                df_sistema.iloc[:, 0]
                .astype(str)
                .str.contains(
                    search,
                    case=False
                )
            )
            |
            (
                df_sistema.iloc[:, 1]
                .astype(str)
                .str.contains(
                    search,
                    case=False
                )
            )
        )

        res = df_sistema[mask]

        for idx in res.index:

            code = clean_code(
                res.iloc[
                    res.index.get_loc(idx),
                    0
                ]
            )

            name = res.iloc[
                res.index.get_loc(idx),
                1
            ]

            if st.button(
                f"{name} ({code})",
                key=f"search_{code}"
            ):

                st.session_state.df_inv = (
                    add_product_to_conteo(
                        df_conteo,
                        code,
                        name
                    )
                )

                st.session_state.selected_code = code
                st.session_state.selected_name = name

                st.rerun()

    # ======================================
    # EDITOR
    # ======================================
    if "selected_code" in st.session_state:

        code = st.session_state.selected_code
        name = st.session_state.selected_name

        match = df_sistema[
            df_sistema.iloc[:, 0]
            .astype(str) == code
        ]

        stock = (
            match.iloc[0, 2]
            if not match.empty
            else 0
        )

        idx = df_conteo[
            df_conteo.iloc[:, 0]
            .astype(str) == code
        ].index[0]

        st.markdown(f"""
        <div class="product-header">
        <b>{name}</b><br>
        Código: {code}<br>
        Stock Sistema: {stock}
        </div>
        """, unsafe_allow_html=True)

        cols = [
            "BO1",
            "BO2",
            "BO3",
            "AL1",
            "AL2",
            "AL3",
            "VALES",
            "VENCIDOS"
        ]

        values = (
            df_conteo.iloc[idx, 3:11]
            .fillna(0)
            .astype(int)
            .tolist()
        )

        while len(values) < 8:
            values.append(0)

        inputs = {}

        row1 = st.columns(4)
        row2 = st.columns(4)

        for i, col_name in enumerate(cols):

            container = (
                row1
                if i < 4
                else row2
            )

            with container[i % 4]:

                inputs[col_name] = (
                    st.number_input(
                        col_name,
                        min_value=0,
                        value=int(values[i]),
                        key=f"{code}_{col_name}"
                    )
                )

        total = sum(inputs.values())

        st.markdown(
            f"<div class='big-font'>TOTAL: {total}</div>",
            unsafe_allow_html=True
        )

        diferencia = total - stock

        st.write(
            f"📊 Diferencia: {diferencia}"
        )

        if st.button(
            "💾 Guardar",
            type="primary"
        ):

            map_cols = {
                "BO1": 3,
                "BO2": 4,
                "BO3": 5,
                "AL1": 6,
                "AL2": 7,
                "AL3": 8,
                "VALES": 9,
                "VENCIDOS": 10
            }

            for k, v in inputs.items():

                df_conteo.iloc[
                    idx,
                    map_cols[k]
                ] = v

            df_conteo.iloc[
                idx,
                11
            ] = total

            st.success(
                "Guardado correctamente"
            )

    # ======================================
    # TABLA
    # ======================================
    st.subheader(
        "📊 Conteo actual"
    )

    st.dataframe(
        df_conteo,
        use_container_width=True
    )

    # ======================================
    # EXPORTAR
    # ======================================
    st.divider()

    if st.button(
        "📥 EXPORTAR RESULTADO"
    ):

        data = save_full_audit(
            df_conteo,
            df_sistema,
            wb
        )

        filename = (
            f"INVENTARIO_"
            f"{sucursal}_"
            f"{fecha}.xlsx"
        )

        st.download_button(
            "Descargar Excel",
            data,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )